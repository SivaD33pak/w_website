"""Seed (or re-seed) the wedding_database.xlsx workbook.

Run from the backend/ directory:

    python -m scripts.seed_database        # fresh seed (overwrites)
    python -m scripts.seed_database --add-guest NAME CODE [FAMILY] [PHONE]
                                           # append a guest without wiping data

Creates four sheets — Guests, RSVP, Events, Gallery — with the exact column
layout documented in ARCHITECTURE.md and populates them with real wedding data.
RSVP starts empty (header only); rows are added by POST /api/rsvp at runtime.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "excel" / "wedding_database.xlsx"

# Column headers per ARCHITECTURE.md. Order matters — ExcelManager reads by index.
GUEST_COLUMNS = ["GuestID", "Name", "Phone", "Family", "InvitationCode", "Status"]
RSVP_COLUMNS = [
    "GuestID",
    "Coming",
    "Adults",
    "Children",
    "FoodPreference",
    "Message",
    "SubmittedAt",
]
EVENT_COLUMNS = ["EventID", "EventName", "Date", "Time", "Venue"]
GALLERY_COLUMNS = ["ImageID", "Title", "ImagePath"]

# ---------------------------------------------------------------------------
# Seed data — the real wedding details for Sushmi & Nijin, 13 Aug 2026
# ---------------------------------------------------------------------------

GUESTS = [
    # GuestID, Name, Phone, Family, InvitationCode, Status
    ("G001", "Sushmi Family", "+91 98470 00001", "Bride", "BRIDE001", "Pending"),
    ("G002", "Nijin Family", "+91 98470 00002", "Groom", "GROOM001", "Pending"),
    ("G003", "Sushmi & Nijin", "+91 98470 00003", "Couple", "LOVE2026", "Pending"),
    ("G004", "Anu & Family", "+91 98470 00004", "Friends", "FRIEND-ANU", "Pending"),
    ("G005", "Thomas & Family", "+91 98470 00005", "Friends", "FRIEND-TOM", "Pending"),
    ("G006", "Relatives - Neyyattinkara", "+91 98470 00006", "Family", "FAM-NEY-01", "Pending"),
    ("G007", "Relatives - Trivandrum", "+91 98470 00007", "Family", "FAM-TVM-01", "Pending"),
    ("G008", "Colleagues - Tech", "+91 98470 00008", "Colleagues", "COL-TECH-01", "Pending"),
]

EVENTS = [
    # EventID, EventName, Date, Time, Venue
    (
        "wedding-ceremony",
        "Holy Mass & Wedding Ceremony",
        "2026-08-13",
        "10:00 AM ONWARDS",
        "Sree Ragam Convention Centre",
    ),
    (
        "reception",
        "Reception",
        "2026-08-13",
        "06:00 PM ONWARDS",
        "Sree Ragam Convention Centre",
    ),
    (
        "dinner",
        "Dinner",
        "2026-08-13",
        "07:30 PM ONWARDS",
        "Sree Ragam Convention Centre",
    ),
]

GALLERY = [
    # ImageID, Title, ImagePath
    (
        "gallery-bride",
        "Bride Portrait",
        "https://storage.googleapis.com/banani-generated-images/generated-images/89485a65-8d3e-42aa-ae78-4c945e71b559.jpg",
    ),
    (
        "gallery-groom",
        "Groom Portrait",
        "https://storage.googleapis.com/banani-generated-images/generated-images/09879492-06cb-4e45-8d15-980194817ffe.jpg",
    ),
    (
        "gallery-couple-garden",
        "Garden Walk",
        "https://storage.googleapis.com/banani-generated-images/generated-images/564d69bb-01c5-4877-8519-088bafef8a.jpg",
    ),
    (
        "gallery-church-interior",
        "Church Interior",
        "https://storage.googleapis.com/banani-generated-images/generated-images/183643c1-0078-4856-891f-33dee7795aac.jpg",
    ),
    (
        "gallery-bouquet",
        "Bouquet",
        "https://storage.googleapis.com/banani-generated-images/generated-images/f7a8d853-9a52-4d6d-a8c7-b3f0e954ea3c.jpg",
    ),
    (
        "gallery-first-kiss",
        "First Kiss",
        "https://storage.googleapis.com/banani-generated-images/generated-images/2daa0de7-8d51-49c9-ad67-a6a85eb8b8f8.jpg",
    ),
    (
        "gallery-flower-field",
        "Flower Field",
        "https://storage.googleapis.com/banani-generated-images/generated-images/8a673d06-31e8-4342-9291-75307ea10417.jpg",
    ),
    (
        "gallery-family",
        "Family Portrait",
        "https://storage.googleapis.com/banani-generated-images/generated-images/50281493-0077-44d2-8080-cf34feaee0ea.jpg",
    ),
]


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def _seed_sheet(ws, columns: list[str], rows: list[tuple]) -> int:
    """Write a header row + data rows. Returns the count of data rows."""
    ws.append(columns)
    for row in rows:
        ws.append(row)
    return len(rows)


def seed_fresh(workbook_path: Path = WORKBOOK_PATH) -> None:
    """Overwrite the workbook with a clean, seeded copy."""
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet so order is deterministic.
    wb.remove(wb.active)

    guest_count = _seed_sheet(wb.create_sheet("Guests"), GUEST_COLUMNS, GUESTS)
    # RSVP starts header-only — filled in by POST /api/rsvp at runtime.
    _seed_sheet(wb.create_sheet("RSVP"), RSVP_COLUMNS, [])
    event_count = _seed_sheet(wb.create_sheet("Events"), EVENT_COLUMNS, EVENTS)
    gallery_count = _seed_sheet(wb.create_sheet("Gallery"), GALLERY_COLUMNS, GALLERY)

    wb.save(workbook_path)
    wb.close()

    print(f"Seeded {workbook_path}")
    print(f"  Guests : {guest_count} rows")
    print(f"  RSVP   : header only (filled at runtime)")
    print(f"  Events : {event_count} rows")
    print(f"  Gallery: {gallery_count} rows")


def add_guest(
    name: str,
    code: str,
    family: str = "Friends",
    phone: str | None = None,
    workbook_path: Path = WORKBOOK_PATH,
) -> None:
    """Append a single guest to the Guests sheet without touching other data."""
    if not workbook_path.exists():
        sys.exit(f"Workbook not found at {workbook_path}. Run a fresh seed first.")

    wb = load_workbook(workbook_path)
    ws = wb["Guests"]

    # Guard against duplicate InvitationCode (case-insensitive).
    header = [c.value for c in ws[1]]
    code_col = header.index("InvitationCode")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[code_col] and str(row[code_col]).strip().lower() == code.lower():
            wb.close()
            sys.exit(f"InvitationCode '{code}' already exists — not added.")

    # Compute next GuestID.
    id_col = header.index("GuestID")
    existing_ids = [
        int(r[id_col][1:]) for r in ws.iter_rows(min_row=2, values_only=True)
        if r[id_col] and str(r[id_col]).startswith("G")
    ]
    next_id = f"G{max(existing_ids, default=0) + 1:03d}"

    ws.append([next_id, name, phone or "", family, code, "Pending"])
    wb.save(workbook_path)
    wb.close()
    print(f"Added guest {next_id} — {name} ({code})")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Seed the wedding Excel database.")
    parser.add_argument(
        "--add-guest",
        nargs="+",
        metavar="VALUE",
        help="Append a guest: NAME CODE [FAMILY] [PHONE]",
    )
    args = parser.parse_args()

    if args.add_guest:
        parts = args.add_guest
        name, code = parts[0], parts[1]
        family = parts[2] if len(parts) > 2 else "Friends"
        phone = parts[3] if len(parts) > 3 else None
        add_guest(name, code, family, phone)
    else:
        seed_fresh()


if __name__ == "__main__":
    main()
