"""ExcelManager — read-only access to the wedding .xlsx workbook.

RSVP writes are handled by GoogleSheetsManager. This class provides fast
in-memory reads for events and other reference data stored in the Excel file.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "excel" / "wedding_database.xlsx"

# Column name → index mappings (0-based) for each sheet.
# These are read from the first header row so they stay in sync with the file.
_EVENT_COLUMNS = ["EventID", "EventName", "Date", "Time", "Venue"]
_GUEST_COLUMNS = ["GuestID", "Name", "Phone", "Family", "InvitationCode", "Status"]


class ExcelManager:
    """Encapsulates read-only access to ``wedding_database.xlsx``.

    All data is loaded into memory at startup for fast access. RSVP writes
    are handled by ``GoogleSheetsManager`` instead.
    """

    def __init__(self, workbook_path: Path = WORKBOOK_PATH) -> None:
        self.workbook_path = workbook_path

        if not workbook_path.exists():
            raise FileNotFoundError(
                f"Excel workbook not found at {workbook_path}"
            )

        # Pre-load sheet caches for fast reads.
        self._events_cache: list[dict[str, str]] = []
        self._guests_cache: list[dict[str, str]] = []
        self._load_caches()

        logger.info(
            "ExcelManager initialised — %d events, %d guests",
            len(self._events_cache),
            len(self._guests_cache),
        )

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _load_caches(self) -> None:
        """Read every sheet into an in-memory list of row-dicts."""
        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)

        self._events_cache = self._sheet_to_dicts(wb["Events"], _EVENT_COLUMNS)
        self._guests_cache = self._sheet_to_dicts(wb["Guests"], _GUEST_COLUMNS)

        wb.close()

    @staticmethod
    def _sheet_to_dicts(
        sheet: Any,
        columns: list[str],
    ) -> list[dict[str, str]]:
        """Convert an openpyxl sheet (with header row) to a list of dicts.

        Empty rows (where the first column is ``None``) are skipped.
        ``datetime`` values are converted to ``YYYY-MM-DD`` strings.
        """
        rows: list[dict[str, str]] = []
        header_passed = False
        for row in sheet.iter_rows(values_only=True):
            if not header_passed:
                header_passed = True
                continue  # skip the header row
            # Skip entirely empty rows.
            if row[0] is None:
                continue
            row_dict: dict[str, str] = {}
            for idx, col_name in enumerate(columns):
                value = row[idx] if idx < len(row) else None
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                row_dict[col_name] = str(value) if value is not None else ""
            rows.append(row_dict)
        return rows

    # ------------------------------------------------------------------
    # READ methods — fast, no lock needed
    # ------------------------------------------------------------------

    def get_events(self) -> list[dict[str, str]]:
        """Return all rows from the Events sheet."""
        return list(self._events_cache)
